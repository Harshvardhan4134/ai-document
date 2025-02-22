import os
import json
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from datetime import datetime

def extract_text_from_pdf(pdf_path):
    text = ""
    with open(pdf_path, "rb") as file:
        reader = PdfReader(file)
        for page in reader.pages:
            extracted_text = page.extract_text()
            if extracted_text:
                text += extracted_text + "\n"
    return text.strip()

def extract_text_from_xlsx(xlsx_path):
    text = []
    try:
        workbook = load_workbook(xlsx_path, data_only=True)
        for sheet in workbook:
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    text.append(row_text)
        return "\n".join(text).strip()
    except Exception as e:
        print(f"❌ Error extracting text from {xlsx_path}: {e}")
        return None

def extract_text_from_json(json_path):
    try:
        with open(json_path, "r", encoding="utf-8") as file:
            data = json.load(file)
            return json.dumps(data, indent=4)
    except Exception as e:
        print(f"❌ Error extracting text from {json_path}: {e}")
        return None

def extract_text_from_folder(folder_path, output_json_path):
    documents = []
    print(f"📂 Scanning folder: {folder_path}")

    for root, _, files in os.walk(folder_path):
        for filename in files:
            file_path = os.path.join(root, filename)
            extracted_text = None

            if filename.endswith(".pdf"):
                extracted_text = extract_text_from_pdf(file_path)
            elif filename.endswith(".xlsx"):
                extracted_text = extract_text_from_xlsx(file_path)
            elif filename.endswith(".json"):
                extracted_text = extract_text_from_json(file_path)
            else:
                print(f"⚠ Skipping unsupported file type: {filename}")
                continue

            if extracted_text:
                file_metadata = {
                    "file_name": filename,
                    "date_uploaded": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "extracted_text": extracted_text,
                    "metadata": {
                        "file_type": filename.split(".")[-1],
                        "file_size": os.path.getsize(file_path)
                    }
                }
                documents.append(file_metadata)
                print(f"✅ Successfully extracted text from {filename}")

    # ✅ **Ensure directory exists before saving JSON**
    output_dir = os.path.dirname(output_json_path)
    os.makedirs(output_dir, exist_ok=True)

    with open(output_json_path, "w", encoding="utf-8") as json_file:
        json.dump(documents, json_file, ensure_ascii=False, indent=4)

    print(f"✅ Extraction complete. {len(documents)} documents saved to {output_json_path}")

# Example Usage
folder_path = "C:/Users/srava/OneDrive/Desktop/Ai model/pumpmanuals"
output_json_path = "C:/Users/srava/Desktop/processed_data.json"
extract_text_from_folder(folder_path, output_json_path)

# def generate_training_data(processed_json, output_file):
#     training_examples = []
    
#     with open(processed_json, "r", encoding="utf-8") as f:
#         documents = json.load(f)
    
#     for doc in documents:
#         extracted_text = doc["extracted_text"]
#         questions = [
#             f"What is the main topic of {doc['file_name']}?",
#             f"Summarize the key points of {doc['file_name']}.",
#             f"What are the important procedures mentioned in {doc['file_name']}?"
#         ]
        
#         for question in questions:
#             training_examples.append({
#                 "messages": [
#                     {"role": "system", "content": "You are an AI trained on company documents."},
#                     {"role": "user", "content": question},
#                     {"role": "assistant", "content": extracted_text[:500]}  # First 500 characters as a sample response
#                 ]
#             })

#     # Save as JSONL file for OpenAI fine-tuning
#     with open(output_file, "w", encoding="utf-8") as out_file:
#         for example in training_examples:
#             out_file.write(json.dumps(example) + "\n")

#     print(f"✅ Training data saved to {output_file}")

# Example usage
# generate_training_data("processed_data.json", "training_data.json")
import json

def generate_training_data(processed_json, output_file):
    training_examples = []

    with open(processed_json, "r", encoding="utf-8") as f:
        documents = json.load(f)

    with open(output_file, "w", encoding="utf-8") as out_file:
        for doc in documents:
            extracted_text = doc["extracted_text"].lower()  # Convert to lowercase
            if len(extracted_text) < 50:
                continue  # Skip very short text

            questions = [
                f"What is the main topic of {doc['file_name']}?",
                f"Summarize the key points of {doc['file_name']}.",
                f"What are the important procedures mentioned in {doc['file_name']}?"
            ]

            for question in questions:
                json_obj = {
                    "prompt": question + "\n\n###",  # Add separator
                    "completion": extracted_text[:500].lower()  # Limit to 500 characters
                }
                out_file.write(json.dumps(json_obj) + "\n")

    print(f"✅ Training data saved to {output_file}")

# Run the function to generate the fixed training data
generate_training_data("processed_data.json", "training_data.jsonl")
